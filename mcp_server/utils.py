import csv
import os
from neo4j import GraphDatabase
from dotenv import load_dotenv
from langchain_text_splitters import RecursiveCharacterTextSplitter

load_dotenv()
import re
import sys

def normalize_name(name: str) -> str:
    """Collapse whitespace and standardize casing so near-duplicate names merge into one node."""
    return re.sub(r'\s+', ' ', name.strip()).title()

from sentence_transformers import SentenceTransformer

_embedder = SentenceTransformer("all-MiniLM-L6-v2")  # loaded once at module import

def embed_texts(texts: list[str]) -> list[list[float]]:
    """Batch-embed a list of strings locally - no API calls, no rate limits."""
    return _embedder.encode(texts).tolist()


def sanitize_relation_type(relation: str) -> str:
    """Convert an LLM-provided relation label into a safe Cypher relationship type."""
    cleaned = re.sub(r'[^A-Za-z0-9_]', '_', relation.strip().upper())
    cleaned = re.sub(r'_+', '_', cleaned).strip('_')
    return cleaned or "RELATED_TO"


class Neo4jSetup:
    def __init__(self, uri=None, username=None, password=None):
        self.uri = uri or os.getenv("NEO4J_URI")
        self.username = username or os.getenv("NEO4J_USERNAME")
        self.password = password or os.getenv("NEO4J_PASSWORD")
        self.driver = None

    def connect(self):
        self.driver = GraphDatabase.driver(self.uri, auth=(self.username, self.password))
        try:
            self.driver.verify_connectivity()
            print("Neo4j connection successful.", file=sys.stderr)
        except Exception as e:
            print(f"Neo4j connection failed: {e}", file=sys.stderr)
            raise

    def close(self):
        """Close the driver and free its connection pool. Call this when
        the server shuts down or the script exits."""
        if self.driver is not None:
            self.driver.close()
            self.driver = None
            print("Neo4j connection closed.", file=sys.stderr)

    def setup_constraints(self):
        with self.driver.session() as session:
            session.run("""
                CREATE CONSTRAINT entity_name_type_unique IF NOT EXISTS
                FOR (e:Entity) REQUIRE (e.name, e.type) IS UNIQUE
            """)
            session.run("""
                CREATE VECTOR INDEX entity_embeddings IF NOT EXISTS
                FOR (e:Entity) ON e.embedding
                OPTIONS {indexConfig: {`vector.dimensions`: 384, `vector.similarity_function`: 'cosine'}}
            """)
            session.run("""
                CREATE VECTOR INDEX relation_type_embeddings IF NOT EXISTS
                FOR (rt:RelationType) ON rt.embedding
                OPTIONS {indexConfig: {`vector.dimensions`: 384, `vector.similarity_function`: 'cosine'}}
            """)
        print("Constraints and vector indexes ensured.", file=sys.stderr)

    def write_graph(self, entities: list[dict], relationships: list[dict]) -> dict:
        if self.driver is None:
            self.connect()

        norm_entities = [
            {**e, "name": normalize_name(e["name"])}
            for e in entities
        ]

        # Embed all entity names in ONE batch call
        if norm_entities:
            names = [e["name"] for e in norm_entities]
            embeddings = embed_texts(names)
            for e, emb in zip(norm_entities, embeddings):
                e["embedding"] = emb

        norm_relationships = []
        for r in relationships:
            norm_relationships.append({
                "source_name": normalize_name(r["source"]["name"]),
                "source_type": r["source"].get("type", "Unknown"),
                "target_name": normalize_name(r["target"]["name"]),
                "target_type": r["target"].get("type", "Unknown"),
                "relation": r["relation"],
            })

        entities_written = 0
        relationships_written = 0
        relationships_skipped = 0

        with self.driver.session() as session:
            if norm_entities:
                try:
                    session.execute_write(self._merge_entities, norm_entities)
                    entities_written = len(norm_entities)
                except Exception as e:
                    print(f"[neo4j] failed to write entities: {e}", file=sys.stderr)
                    # Entities failed to write - relationships referencing them
                    # would just re-create bare nodes via MERGE anyway, but skip
                    # the relationship step here since something is clearly wrong
                    # with this batch (e.g. bad embedding shape, connection drop).
                    return {
                        "entities_written": 0,
                        "relationships_written": 0,
                        "relationships_skipped": len(norm_relationships),
                    }

            by_type: dict[str, list[dict]] = {}
            for rel in norm_relationships:
                rel_type = sanitize_relation_type(rel["relation"])
                by_type.setdefault(rel_type, []).append(rel)

            # Embed all distinct relation types in ONE batch call
            if by_type:
                rel_type_names = list(by_type.keys())
                rel_type_embeddings = embed_texts(rel_type_names)
                session.execute_write(
                    self._merge_relation_types,
                    list(zip(rel_type_names, rel_type_embeddings))
                )

            for rel_type, rels in by_type.items():
                try:
                    session.execute_write(self._merge_relationships, rel_type, rels)
                    relationships_written += len(rels)
                except Exception as e:
                    print(f"[neo4j] failed to write relation type {rel_type}: {e}", file=sys.stderr)
                    relationships_skipped += len(rels)

        return {
            "entities_written": entities_written,
            "relationships_written": relationships_written,
            "relationships_skipped": relationships_skipped,
        }

    @staticmethod
    def _merge_entities(tx, entities: list[dict]):
        # Only overwrite an existing description when the incoming one is
        # non-empty - an empty string from a later chunk should never wipe
        # out a good description written by an earlier chunk.
        tx.run(
            """
            UNWIND $entities AS e
            MERGE (n:Entity {name: e.name, type: e.type})
            SET n.description = CASE
                    WHEN e.description IS NOT NULL AND e.description <> ''
                    THEN e.description
                    ELSE coalesce(n.description, '')
                END,
                n.embedding = e.embedding
            """,
            entities=entities,
        )

    @staticmethod
    def _merge_relation_types(tx, type_embeddings: list[tuple]):
        tx.run("""
            UNWIND $items AS item
            MERGE (rt:RelationType {name: item[0]})
            ON CREATE SET rt.embedding = item[1]
            """, items=[list(t) for t in type_embeddings])

    @staticmethod
    def _merge_relationships(tx, rel_type: str, rels: list[dict]):
        # rel_type is sanitized (alnum/underscore only) before reaching here — safe to interpolate.
        # MERGE on the relationship itself (not just the endpoints) so the same
        # fact extracted multiple times - across chunks, or across separate runs -
        # collapses into a single edge instead of piling up duplicates.
        query = f"""
            UNWIND $rels AS r
            MERGE (a:Entity {{name: r.source_name, type: r.source_type}})
            MERGE (b:Entity {{name: r.target_name, type: r.target_type}})
            MERGE (a)-[rel:`{rel_type}`]->(b)
        """
        tx.run(query, rels=rels)

    def _resolve_entity(self, session, entity_query: str):
        """Vector-search a fuzzy entity mention to a real (name, type, description) node."""
        embedding = embed_texts([entity_query])[0]
        match = session.run("""
            CALL db.index.vector.queryNodes('entity_embeddings', 1, $embedding)
            YIELD node, score
            RETURN node.name AS name, node.type AS type,
                   node.description AS description, score
        """, embedding=embedding).single()
        if match is None or match["score"] < 0.3:
            return None
        return match

    def _resolve_relation(self, session, relation_query: str):
        """Vector-search a fuzzy relationship phrase to a real relationship type."""
        embedding = embed_texts([relation_query])[0]
        match = session.run("""
            CALL db.index.vector.queryNodes('relation_type_embeddings', 1, $embedding)
            YIELD node, score
            RETURN node.name AS name, score
        """, embedding=embedding).single()
        if match is None or match["score"] < 0.3:
            return None
        return match["name"]

    def get_entity_context(
        self,
        entity_query: str,
        relation_query: str = None,
        target_entity_query: str = None,
        limit: int = 20,
    ) -> dict:
        if self.driver is None:
            self.connect()

        with self.driver.session() as session:
            # Step 1: resolve the primary entity
            match = self._resolve_entity(session, entity_query)
            print(f"[get_entity_context] query={entity_query!r} match={match}", file=sys.stderr)
            if match is None:
                return {"found": False, "query": entity_query}

            resolved_name = match["name"]
            resolved_type = match["type"]

            # Step 2: resolve the relation phrase, if given
            relation_filter = None
            if relation_query:
                relation_filter = self._resolve_relation(session, relation_query)

            # Step 3a: TWO-ENTITY mode - relationship(s) between two specific entities
            if target_entity_query:
                target_match = self._resolve_entity(session, target_entity_query)
                print(f"[get_entity_context] target_query={target_entity_query!r} match={target_match}", file=sys.stderr)
                if target_match is None:
                    return {
                        "found": False,
                        "query": entity_query,
                        "target_query": target_entity_query,
                        "reason": "target entity not found in graph",
                    }

                target_name = target_match["name"]
                target_type = target_match["type"]

                if relation_filter:
                    rel_result = session.run(f"""
                        MATCH (a:Entity {{name: $a_name, type: $a_type}})
                              -[r:`{relation_filter}`]-
                              (b:Entity {{name: $b_name, type: $b_type}})
                        RETURN type(r) AS relation
                        LIMIT $limit
                    """, a_name=resolved_name, a_type=resolved_type,
                         b_name=target_name, b_type=target_type, limit=limit)
                else:
                    rel_result = session.run("""
                        MATCH (a:Entity {name: $a_name, type: $a_type})
                              -[r]-
                              (b:Entity {name: $b_name, type: $b_type})
                        RETURN type(r) AS relation
                        LIMIT $limit
                    """, a_name=resolved_name, a_type=resolved_type,
                         b_name=target_name, b_type=target_type, limit=limit)

                relations = [r["relation"] for r in rel_result]

                return {
                    "found": len(relations) > 0,
                    "entity_name": resolved_name,
                    "entity_type": resolved_type,
                    "entity_description": match["description"],
                    "target_entity_name": target_name,
                    "target_entity_type": target_type,
                    "target_entity_description": target_match["description"],
                    "resolved_relation": relation_filter,
                    "relations": relations,
                }

            # Step 3b: SINGLE-ENTITY mode - all connections, optionally filtered by relation
            if relation_filter:
                conn_result = session.run(f"""
                    MATCH (e:Entity {{name: $name, type: $type}})-[r:`{relation_filter}`]-(other:Entity)
                    RETURN type(r) AS relation, other.name AS connected_name,
                           other.type AS connected_type, other.description AS connected_description
                    LIMIT $limit
                """, name=resolved_name, type=resolved_type, limit=limit)
            else:
                conn_result = session.run("""
                    MATCH (e:Entity {name: $name, type: $type})-[r]-(other:Entity)
                    RETURN type(r) AS relation, other.name AS connected_name,
                           other.type AS connected_type, other.description AS connected_description
                    LIMIT $limit
                """, name=resolved_name, type=resolved_type, limit=limit)

            connections = [dict(r) for r in conn_result]

        return {
            "found": True,
            "entity_name": resolved_name,
            "entity_type": resolved_type,
            "entity_description": match["description"],
            "resolved_relation": relation_filter,
            "connections": connections,
        }


from typing import Iterator
from pypdf import PdfReader
from docx import Document
from pathlib import Path
from openpyxl import load_workbook



def stream_pages(file_path: str) -> Iterator[str]:
    """Yield text page-by-page (PDF), paragraph-block-by-block (DOCX), or
    row-block-by-block (XLSX) instead of loading the whole document into
    one string."""
    path = Path(file_path)
    if path.suffix == ".pdf":
        reader = PdfReader(str(path))
        for page in reader.pages:
            yield page.extract_text() or ""

    elif path.suffix == ".docx":
        document = Document(str(path))
        buf = []
        for p in document.paragraphs:
            buf.append(p.text)
            if len("\n".join(buf)) > 2000:
                yield "\n".join(buf)
                buf = []
        if buf:
            yield "\n".join(buf)

    elif path.suffix in (".xlsx", ".xlsm"):
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                continue  # empty sheet

            headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(headers)]
            buf = [f"Sheet: {sheet.title}"]

            for row in rows_iter:
                if row is None or all(v is None for v in row):
                    continue
                row_text = ", ".join(
                    f"{h}: {v}" for h, v in zip(headers, row) if v is not None
                )
                buf.append(row_text)
                if len("\n".join(buf)) > 2000:
                    yield "\n".join(buf)
                    buf = [f"Sheet: {sheet.title} (continued)"]

            if len(buf) > 1:  # more than just the header line
                yield "\n".join(buf)
    elif path.suffix == ".csv":
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            try:
                headers = next(reader)
            except StopIteration:
                return  # empty file

            buf = []
            for row in reader:
                if not row or all(v == "" for v in row):
                    continue
                row_text = ", ".join(
                    f"{h}: {v}" for h, v in zip(headers, row) if v != ""
                )
                buf.append(row_text)
                if len("\n".join(buf)) > 2000:
                    yield "\n".join(buf)
                    buf = []
            if buf:
                yield "\n".join(buf)
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")

def chunk_document_streaming(
    file_path: str,
    chunk_size: int = 1000,
    chunk_overlap: int = 150,
) -> Iterator[str]:
    """Stream chunks from a large document without holding the full text
    or full chunk list in memory at once. Suitable for 1000s of pages."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    buffer = ""
    # keep buffer roughly 3x chunk_size before splitting, so the splitter
    # has enough context to find good paragraph/sentence boundaries
    flush_threshold = chunk_size * 3

    for page_text in stream_pages(file_path):
        buffer += ("\n\n" + page_text if buffer else page_text)

        if len(buffer) >= flush_threshold:
            pieces = splitter.split_text(buffer)
            # hold back the last piece — it may be incomplete without
            # the next page's text, so carry it into the next round
            for piece in pieces[:-1]:
                yield piece
            buffer = pieces[-1] if pieces else ""

    # flush whatever's left after the last page
    if buffer:
        for piece in splitter.split_text(buffer):
            yield piece